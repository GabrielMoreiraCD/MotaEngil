"""
bom_generator.py — Etapa 5: Geração da Lista de Materiais (BOM)
=================================================================
Recebe MateriaisRequisitados (Etapa 4) + EscopoTriagemUnificado (Etapa 2) e:
  1. Deduplica itens com mesmo codigo_material (soma quantidades)
  2. Atribui número sequencial a cada item
  3. Gera advertências automáticas (itens sem código, gaps de catálogo)
  4. Exporta para XLSX (openpyxl) e/ou PDF (reportlab)

Saída: ListaMateriais (persistida em etapa5_bom_result.json)
       + arquivo XLSX/PDF na pasta de saída
"""

import json
import logging
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from core.schemas import (
    BOMLineItem,
    CatalogItemMatch,
    EscopoTriagemUnificado,
    ListaMateriais,
    MateriaisRequisitados,
)

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# BOMGenerator
# ─────────────────────────────────────────────────────────────────────────────

class BOMGenerator:
    """
    Consolida os itens mapeados na Etapa 4 em uma Lista de Materiais estruturada
    e exporta para XLSX e/ou PDF.
    """

    def generate(
        self,
        materiais: MateriaisRequisitados,
        escopo: EscopoTriagemUnificado,
        normas_base: list[str],
    ) -> ListaMateriais:
        """
        Gera a ListaMateriais a partir dos itens mapeados.

        Args:
            materiais: Saída da Etapa 4 (itens do catálogo)
            escopo: Dados do serviço (header da BOM)
            normas_base: Lista de normas consultadas na Etapa 3
        """
        deduped = self._deduplicate(materiais.itens)
        bom_items = self._build_line_items(deduped)
        advertencias = self._build_advertencias(materiais, escopo)

        itens_sem_codigo = sum(1 for i in bom_items if not i.codigo_material)

        bom = ListaMateriais(
            id_servico=escopo.id_servico,
            titulo_servico=escopo.titulo_servico,
            plataforma=escopo.plataforma,
            tag_equipamento=escopo.tag_linha_principal or escopo.tag_equipamento_principal,
            data_geracao=datetime.now(timezone.utc).isoformat(),
            itens=bom_items,
            total_itens=len(bom_items),
            itens_sem_codigo=itens_sem_codigo,
            advertencias=advertencias,
            normas_base=normas_base,
        )

        log.info(
            f"[Etapa 5] BOM gerada: {len(bom_items)} itens "
            f"({itens_sem_codigo} sem código de catálogo)"
        )
        return bom

    # ── Deduplication ─────────────────────────────────────────────────────────

    def _deduplicate(self, itens: list[CatalogItemMatch]) -> list[CatalogItemMatch]:
        """
        Agrupa itens com o mesmo codigo_material e soma quantidades.
        Itens sem código (unmapped) são mantidos individualmente pois
        representam requisitos distintos não resolvidos.
        """
        # Itens com código → agrupa
        by_code: dict[str, list[CatalogItemMatch]] = defaultdict(list)
        unmapped: list[CatalogItemMatch] = []

        for item in itens:
            if item.codigo and item.mapeamento_status != "unmapped":
                by_code[item.codigo].append(item)
            else:
                unmapped.append(item)

        deduped: list[CatalogItemMatch] = []

        for codigo, grupo in by_code.items():
            representante = grupo[0]
            qtd_total = sum(i.quantidade for i in grupo)
            merged = representante.model_copy(update={"quantidade": qtd_total})
            deduped.append(merged)

        deduped.extend(unmapped)
        return deduped

    # ── Line items ────────────────────────────────────────────────────────────

    def _build_line_items(self, itens: list[CatalogItemMatch]) -> list[BOMLineItem]:
        """Converte CatalogItemMatch para BOMLineItem com número sequencial."""
        line_items: list[BOMLineItem] = []

        for i, item in enumerate(itens, start=1):
            # Especificação técnica consolidada
            specs = [
                s for s in [
                    item.especificacao_catalogo,
                    item.norma_catalogo,
                    item.material_base,
                ]
                if s
            ]
            spec_str = " | ".join(specs) if specs else item.requisito_origem

            observacoes = item.observacoes
            if item.mapeamento_status == "partial":
                nota = f"Mapeamento parcial (score={item.score_similaridade:.2f}). Verificar especificação."
                observacoes = (observacoes + " | " + nota) if observacoes else nota

            line_items.append(BOMLineItem(
                item_numero=i,
                codigo_material=item.codigo,
                descricao=item.descricao_catalogo,
                especificacao_tecnica=spec_str,
                diametro_nps=item.diametro,
                quantidade=item.quantidade,
                unidade=item.unidade_fornecimento,
                quantidade_estimada=item.quantidade_estimada,
                norma_origem="",  # será preenchido pelo caller se necessário
                categoria=item.categoria_catalogo,
                source_file_catalogo=item.source_file,
                observacoes=observacoes,
            ))

        return line_items

    # ── Advertências ──────────────────────────────────────────────────────────

    def _build_advertencias(
        self,
        materiais: MateriaisRequisitados,
        escopo: EscopoTriagemUnificado,
    ) -> list[str]:
        """Gera lista de avisos para o engenheiro revisor."""
        advertencias: list[str] = []

        if materiais.total_nao_mapeados > 0:
            # Coleta categorias não mapeadas
            unmapped_descricoes = [
                i.descricao_catalogo
                for i in materiais.itens
                if i.mapeamento_status == "unmapped"
            ]
            advertencias.append(
                f"{materiais.total_nao_mapeados} item(ns) não encontrado(s) no catálogo: "
                + "; ".join(unmapped_descricoes[:5])
                + (f" ... (e mais {len(unmapped_descricoes) - 5})" if len(unmapped_descricoes) > 5 else "")
                + ". Verificar catálogo atualizado ou adicionar planilhas de consumíveis."
            )

        partial_count = sum(1 for i in materiais.itens if i.mapeamento_status == "partial")
        if partial_count > 0:
            advertencias.append(
                f"{partial_count} item(ns) com mapeamento parcial (score baixo). "
                "Revisar especificação técnica antes de emitir requisição de compra."
            )

        qty_estimada = sum(1 for i in materiais.itens if i.quantidade_estimada)
        if qty_estimada > 0:
            advertencias.append(
                f"Quantidades de {qty_estimada} item(ns) foram ESTIMADAS (padrão: 1 UN). "
                "Verificar isométrico/memorial descritivo para quantidades reais antes de comprar."
            )

        if not escopo.materiais_criticos:
            advertencias.append(
                "Nenhum material crítico listado no formulário de serviço (campo 4.8). "
                "Confirmar se lista de materiais críticos está completa."
            )

        return advertencias

    # ── Export XLSX ───────────────────────────────────────────────────────────

    def export_xlsx(self, bom: ListaMateriais, output_path: Path) -> Path:
        """
        Exporta a BOM para planilha Excel (.xlsx).
        Usa openpyxl com formatação básica para uso em compras.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError as e:
            raise ImportError("openpyxl não encontrado. Execute: pip install openpyxl>=3.1.0") from e

        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Materiais"

        # ── Cabeçalho do documento ────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="003366")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(bold=True, size=12)
        warn_fill = PatternFill("solid", fgColor="FFF3CD")
        warn_font = Font(color="856404", size=9)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row = 1
        ws.cell(row=row, column=1, value="LISTA DE MATERIAIS — PIPELINE AUTOMÁTICO").font = Font(bold=True, size=14)
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        meta_fields = [
            ("Serviço:", bom.id_servico),
            ("Título:", bom.titulo_servico),
            ("Plataforma:", bom.plataforma),
            ("TAG Equipamento:", bom.tag_equipamento),
            ("Data Geração:", bom.data_geracao[:19].replace("T", " ")),
            ("Gerado por:", bom.gerado_por),
            ("Normas Base:", ", ".join(bom.normas_base)),
        ]
        for label, value in meta_fields:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
            ws.cell(row=row, column=2, value=value)
            ws.merge_cells(f"B{row}:J{row}")
            row += 1

        row += 1

        # ── Advertências ──────────────────────────────────────────────────────
        if bom.advertencias:
            ws.cell(row=row, column=1, value="⚠ ADVERTÊNCIAS").font = Font(bold=True, color="856404")
            ws.merge_cells(f"A{row}:J{row}")
            row += 1
            for adv in bom.advertencias:
                cell = ws.cell(row=row, column=1, value=adv)
                cell.font = warn_font
                cell.fill = warn_fill
                ws.merge_cells(f"A{row}:J{row}")
                row += 1
            row += 1

        # ── Tabela de itens ───────────────────────────────────────────────────
        col_headers = [
            "Item", "Código", "Descrição", "Especificação Técnica",
            "DN (pol.)", "Qtd.", "Unid.", "Qtd. Est.?",
            "Categoria", "Observações",
        ]
        col_widths = [6, 18, 40, 35, 10, 8, 8, 10, 25, 40]

        for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        row += 1
        header_row = row - 1

        # Linhas de dados
        alt_fill = PatternFill("solid", fgColor="EEF2FF")
        for item in bom.itens:
            fill = alt_fill if item.item_numero % 2 == 0 else PatternFill()
            row_data = [
                item.item_numero,
                item.codigo_material or "",
                item.descricao,
                item.especificacao_tecnica,
                item.diametro_nps or "",
                item.quantidade,
                item.unidade,
                "SIM" if item.quantidade_estimada else "NÃO",
                item.categoria,
                item.observacoes or "",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_idx in (6,):  # Qtd. → número
                    cell.number_format = "0.00"
            row += 1

        # ── Linha de totais ───────────────────────────────────────────────────
        ws.cell(row=row, column=1, value="TOTAIS").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{bom.total_itens} itens").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"{bom.itens_sem_codigo} sem código").font = Font(bold=True, color="CC0000")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        log.info(f"  XLSX exportado: {output_path}")
        return output_path

    # ── Export PDF ────────────────────────────────────────────────────────────

    def export_pdf(self, bom: ListaMateriais, output_path: Path) -> Path:
        """
        Exporta a BOM para PDF usando reportlab.
        Formato A4 paisagem com tabela de materiais.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
        except ImportError as e:
            raise ImportError("reportlab não encontrado. Execute: pip install reportlab>=4.0.0") from e

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=14, spaceAfter=6)
        meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, spaceAfter=2)
        warn_style = ParagraphStyle("warn", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#856404"), spaceAfter=2)
        cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, wordWrap="CJK")

        story = []

        # Título
        story.append(Paragraph("LISTA DE MATERIAIS — PIPELINE AUTOMÁTICO", title_style))
        story.append(Paragraph(f"<b>Serviço:</b> {bom.id_servico} | <b>Plataforma:</b> {bom.plataforma} | "
                               f"<b>TAG:</b> {bom.tag_equipamento}", meta_style))
        story.append(Paragraph(f"<b>Título:</b> {bom.titulo_servico}", meta_style))
        story.append(Paragraph(f"<b>Data geração:</b> {bom.data_geracao[:19].replace('T', ' ')} | "
                               f"<b>Normas:</b> {', '.join(bom.normas_base)}", meta_style))
        story.append(Spacer(1, 4 * mm))

        if bom.advertencias:
            story.append(Paragraph("<b>⚠ ADVERTÊNCIAS</b>", warn_style))
            for adv in bom.advertencias:
                story.append(Paragraph(f"• {adv}", warn_style))
            story.append(Spacer(1, 4 * mm))

        # Tabela
        col_headers = ["Item", "Código", "Descrição", "Especificação", "DN", "Qtd.", "Unid.", "Categoria", "Observações"]
        col_widths_mm = [10, 25, 65, 55, 12, 12, 12, 40, 55]

        table_data = [[Paragraph(f"<b>{h}</b>", cell_style) for h in col_headers]]

        for item in bom.itens:
            row = [
                Paragraph(str(item.item_numero), cell_style),
                Paragraph(item.codigo_material or "—", cell_style),
                Paragraph(item.descricao, cell_style),
                Paragraph(item.especificacao_tecnica, cell_style),
                Paragraph(item.diametro_nps or "—", cell_style),
                Paragraph(f"{item.quantidade:.1f}", cell_style),
                Paragraph(item.unidade, cell_style),
                Paragraph(item.categoria, cell_style),
                Paragraph(item.observacoes or "", cell_style),
            ]
            table_data.append(row)

        tbl = Table(
            table_data,
            colWidths=[w * mm for w in col_widths_mm],
            repeatRows=1,
        )
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#AAAAAA")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(tbl)

        doc.build(story)
        log.info(f"  PDF exportado: {output_path}")
        return output_path

    # ── Persist JSON ─────────────────────────────────────────────────────────

    def save_json(self, bom: ListaMateriais, output_path: Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(bom.model_dump(), f, indent=2, ensure_ascii=False)
        log.info(f"  JSON exportado: {output_path}")
        return output_path
